# SPDX-FileCopyrightText: Copyright (c) 2025 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import excel2img
import numpy as np

import argparse


def populate(ws, data, cells):

    nil = 0
    csirs = 1
    ssb = 3
    pdcch = 4
    pdsch = 5
    pusch = 6
    pucch = 7
    prach = 8
    missing = -1

    color_mapping = {}
    color_mapping[nil] = "6F6F6F"
    color_mapping[csirs] = "8EA9DB"
    color_mapping[ssb] = "FFC000"
    color_mapping[pdcch] = "70AD47"
    color_mapping[pdsch] = "9933FF"
    color_mapping[nil] = "6f6f6f"
    color_mapping[pucch] = "e404ba"
    color_mapping[prach] = "ddd5be"
    color_mapping[pusch] = "3d85c6"
    color_mapping[missing] = "ff0000"
    color_mapping[missing] = "FF0000"

    border = Border(left=Side(style="thick"), right=Side(style="thick"))

    total_prbs = 273
    total_syms = 14

    sequence = data["sequence"]
    types = data["types"]

    off_rows = 1
    off_cols = 1

    for c_idx in range(cells):
        for k in range(total_prbs):
            ws.row_dimensions[
                (c_idx + 1) * total_prbs - k + off_rows + c_idx
            ].height = 1.5

    for slt_idx, slot in enumerate(sequence):

        for c_idx in range(cells):

            if "dl" in slot:

                color = color_mapping[nil]

                for s in range(total_syms):

                    for k in range(total_prbs):

                        ws.cell(
                            row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                            column=slt_idx * total_syms + s + off_cols + slt_idx,
                        ).fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

                        ws.cell(
                            row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                            column=slt_idx * total_syms + s + off_cols + slt_idx,
                        ).border = border

                pdcch_sym_start = types[slot].get("pdcch_sym_start", None)
                pdcch_sym_end = types[slot].get("pdcch_sym_end", None)
                pdcch_prb = types[slot].get("pdcch_prb", None)

                if (
                    pdcch_sym_start is not None
                    and pdcch_sym_end is not None
                    and pdcch_prb is not None
                ):

                    color = color_mapping[pdcch]

                    for idx in range(len(pdcch_sym_start)):

                        for s in range(pdcch_sym_start[idx], pdcch_sym_end[idx] + 1):

                            for k in range(pdcch_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

                ssb_sym_start = types[slot].get("ssb_sym_start", None)
                ssb_sym_end = types[slot].get("ssb_sym_end", None)
                ssb_prb = types[slot].get("ssb_prb", None)

                if (
                    ssb_sym_start is not None
                    and ssb_sym_end is not None
                    and ssb_prb is not None
                ):

                    color = color_mapping[ssb]

                    for idx in range(len(ssb_sym_start)):

                        for s in range(ssb_sym_start[idx], ssb_sym_end[idx] + 1):

                            for k in range(ssb_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

                pdsch_sym_start = types[slot].get("pdsch_sym_start", None)
                pdsch_sym_end = types[slot].get("pdsch_sym_end", None)
                pdsch_prb = types[slot].get("pdsch_prb", None)

                if (
                    pdsch_sym_start is not None
                    and pdsch_sym_end is not None
                    and pdsch_prb is not None
                ):

                    color = color_mapping[pdsch]

                    off_ue = 0

                    if (
                        ssb_sym_start is not None
                        and ssb_sym_end is not None
                        and ssb_prb is not None
                    ):
                        off_ue += ssb_prb

                    for idx in range(len(pdsch_sym_start)):

                        for s in range(pdsch_sym_start[idx], pdsch_sym_end[idx] + 1):

                            for k in range(pdsch_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs
                                    - k
                                    - off_ue
                                    + off_rows
                                    + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

                        off_ue += pdsch_prb

                csirs_flex_sym_start = types[slot].get("csirs_flex_sym_start", None)
                csirs_flex_sym_end = types[slot].get("csirs_flex_sym_end", None)

                if csirs_flex_sym_start is not None and csirs_flex_sym_end is not None:

                    color = color_mapping[csirs]

                    toggle = 0

                    for idx in range(len(csirs_flex_sym_start)):

                        if c_idx % 2 == toggle:

                            for s in range(
                                csirs_flex_sym_start[idx], csirs_flex_sym_end[idx] + 1
                            ):

                                off_csirs = 0

                                if (
                                    ssb_sym_start is not None
                                    and ssb_sym_end is not None
                                    and ssb_prb is not None
                                ):

                                    ssb_syms = []

                                    for ssb_idx, ssb_itm in enumerate(ssb_sym_start):
                                        ssb_syms.extend(
                                            list(
                                                np.arange(
                                                    ssb_sym_start[ssb_idx],
                                                    ssb_sym_end[ssb_idx] + 1,
                                                )
                                            )
                                        )

                                    if s in ssb_syms:
                                        off_csirs += ssb_prb

                                for k in range(off_csirs, total_prbs):

                                    ws.cell(
                                        row=(c_idx + 1) * total_prbs
                                        - k
                                        + off_rows
                                        + c_idx,
                                        column=slt_idx * total_syms
                                        + s
                                        + off_cols
                                        + slt_idx,
                                    ).fill = PatternFill(
                                        start_color=color,
                                        end_color=color,
                                        fill_type="solid",
                                    )

                        toggle = abs(toggle - 1)

                csirs_fix_sym_start = types[slot].get("csirs_fix_sym_start", None)
                csirs_fix_sym_end = types[slot].get("csirs_fix_sym_end", None)

                if csirs_fix_sym_start is not None and csirs_fix_sym_end is not None:

                    color = color_mapping[csirs]

                    for idx in range(len(csirs_fix_sym_start)):

                        for s in range(
                            csirs_fix_sym_start[idx], csirs_fix_sym_end[idx] + 1
                        ):

                            off_csirs = 0

                            if (
                                ssb_sym_start is not None
                                and ssb_sym_end is not None
                                and ssb_prb is not None
                            ):
                                off_csirs += ssb_prb

                            for k in range(off_csirs, total_prbs):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

            elif "ul" in slot:

                color = color_mapping[nil]

                for s in range(total_syms):

                    for k in range(total_prbs):

                        ws.cell(
                            row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                            column=slt_idx * total_syms + s + off_cols + slt_idx,
                        ).fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

                        ws.cell(
                            row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                            column=slt_idx * total_syms + s + off_cols + slt_idx,
                        ).border = border

                pucch_sym_start = types[slot].get("pucch_sym_start", None)
                pucch_sym_end = types[slot].get("pucch_sym_end", None)
                pucch_prb = types[slot].get("pucch_prb", None)

                if (
                    pucch_sym_start is not None
                    and pucch_sym_end is not None
                    and pucch_prb is not None
                ):

                    color = color_mapping[pucch]

                    for idx in range(len(pucch_sym_start)):

                        for s in range(pucch_sym_start[idx], pucch_sym_end[idx] + 1):

                            for k in range(pucch_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs - k + off_rows + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

                pusch_sym_start = types[slot].get("pusch_sym_start", None)
                pusch_sym_end = types[slot].get("pusch_sym_end", None)
                pusch_prb = types[slot].get("pusch_prb", None)

                if (
                    pusch_sym_start is not None
                    and pusch_sym_end is not None
                    and pusch_prb is not None
                ):

                    color = color_mapping[pusch]

                    off_ue = 0

                    if (
                        pucch_sym_start is not None
                        and pucch_sym_end is not None
                        and pucch_prb is not None
                    ):
                        off_ue += pucch_prb

                    for idx in range(len(pusch_sym_start)):

                        for s in range(pusch_sym_start[idx], pusch_sym_end[idx] + 1):

                            for k in range(pusch_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs
                                    - k
                                    - off_ue
                                    + off_rows
                                    + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )

                        off_ue += pusch_prb

                prach_sym_start = types[slot].get("prach_sym_start", None)
                prach_sym_end = types[slot].get("prach_sym_end", None)
                prach_prb = types[slot].get("prach_prb", None)

                if (
                    prach_sym_start is not None
                    and prach_sym_end is not None
                    and prach_prb is not None
                ):

                    color = color_mapping[prach]

                    off_prach = 0

                    if (
                        pucch_sym_start is not None
                        and pucch_sym_end is not None
                        and pucch_prb is not None
                    ):
                        off_prach += pucch_prb

                    if (
                        pusch_sym_start is not None
                        and pusch_sym_end is not None
                        and pusch_prb is not None
                    ):
                        off_prach += pusch_prb * len(pusch_sym_start)

                    for idx in range(len(prach_sym_start)):

                        for s in range(prach_sym_start[idx], prach_sym_end[idx] + 1):

                            for k in range(prach_prb):

                                ws.cell(
                                    row=(c_idx + 1) * total_prbs
                                    - k
                                    - off_prach
                                    + off_rows
                                    + c_idx,
                                    column=slt_idx * total_syms
                                    + s
                                    + off_cols
                                    + slt_idx,
                                ).fill = PatternFill(
                                    start_color=color,
                                    end_color=color,
                                    fill_type="solid",
                                )


base = argparse.ArgumentParser()
base.add_argument(
    "--config",
    type=str,
    nargs="+",
    dest="config",
    help="Specifies the configuration file",
    required=True,
)
args = base.parse_args()


for config in args.config:

    ifile = open(config, "r")
    data = json.load(ifile)
    ifile.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Pattern"

    cells = 16

    populate(ws, data, cells)

    wb.save(config.replace(".json", ".xlsx"))

    excel2img.export_img(
        config.replace(".json", ".xlsx"),
        config.replace(".json", ".png"),
        "Pattern",
    )
